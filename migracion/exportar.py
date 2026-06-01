# Extrae los clientes antiguos del Excel a clientes-antiguos.json.
# Uso:  python migracion/exportar.py "ruta/al/Excel.xlsx"
import openpyxl, json, re, sys, os

EXCEL = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\alekc\Downloads\Musicala Vacacional 2025 .xlsx"
SALIDA = os.path.join(os.path.dirname(__file__), "clientes-antiguos.json")

def limpiar(v):
    if v is None: return ""
    s = str(v).strip()
    return s[:-2] if re.fullmatch(r"\d+\.0", s) else s

wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
contactos, vistos = [], set()

# Interesados pendientes (pipeline histórico)
if "Pendientes interesados 2025" in wb.sheetnames:
    for r in list(wb["Pendientes interesados 2025"].iter_rows(values_only=True))[1:]:
        est, acu, cel = limpiar(r[0]), limpiar(r[1]), limpiar(r[2])
        com = limpiar(r[7]) if len(r) > 7 else ""
        if not (est or acu or cel): continue
        key = (est.lower(), cel)
        if key in vistos: continue
        vistos.add(key)
        contactos.append({"estudiante": est, "acudiente": acu, "celular": cel,
                          "comentario": com, "estado": "Antiguo", "origen": "Antiguo"})

# Inscritos históricos de varias temporadas
for hoja in ["Estudiantes inscritos Diciembre", "062025 inscritos Vacacionales",
             "Estudiantes inscritos 0624", "Estudiantes inscritos 1223", "Estudiantes inscritos 0623"]:
    if hoja not in wb.sheetnames: continue
    rows = list(wb[hoja].iter_rows(values_only=True))
    if not rows: continue
    hdr = [limpiar(c).lower() for c in rows[0]]
    def idx(name): return next((i for i, h in enumerate(hdr) if name in h), None)
    iEst, iEdad, iGrupo = idx("estudiante"), idx("edad"), idx("grupo")
    for r in rows[1:]:
        est = limpiar(r[iEst]) if iEst is not None and iEst < len(r) else ""
        if not est or (est.lower(), "") in vistos: continue
        vistos.add((est.lower(), ""))
        c = {"estudiante": est, "estado": "Antiguo", "origen": "Antiguo"}
        edad = limpiar(r[iEdad]) if iEdad is not None and iEdad < len(r) else ""
        grupo = limpiar(r[iGrupo]) if iGrupo is not None and iGrupo < len(r) else ""
        if edad:
            try: c["edad"] = int(float(edad))
            except ValueError: pass
        if grupo: c["grupo"] = grupo.strip()
        contactos.append(c)

with open(SALIDA, "w", encoding="utf-8") as f:
    json.dump(contactos, f, ensure_ascii=False, indent=1)
print(f"{len(contactos)} contactos -> {SALIDA}")
